import logging
import os.path
import pandas as pd

from openpyxl import Workbook, load_workbook, styles
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime


def load_wb(filename):
    # Make sure the excel sheet is not open
    i = 0
    while True:
        if i > 5:
            raise RuntimeError("You tried too many times")
        try:
            f = open(filename, 'r+')
            break
        except IOError:
            input("Please close the open file in Excel and press enter. ")
            logging.warning("User tried to open a file that is locked and is likely open in Excel.")
            i += 1

    return load_workbook(filename)


def save_wb(filename, wb):
    try:
        wb.save(filename)
    except PermissionError as E:
        logging.warning("Permission error when saving.\n{}".format(E))
        filename += "Conflicted Copy.xlsx"
        wb.save(filename)


def next_table_name(wb):
    tables = {}
    for sheet in wb.sheetnames:
        tables = tables | wb[sheet].tables
    base = 1
    while True:
        if base > 100:
            raise RuntimeError("Too many tables")
        if "Table{}".format(base) in tables:
            base += 1
        else:
            return "Table{}".format(base)


def dataframe(form_fields, form_type):
    return form_fields[form_type]["dataframe"]


def write_row(ws_tab_name, data, filename):
    # save the workbook
    wb = load_wb(filename)
    ws = wb[ws_tab_name]
    ws.append(data)
    save_wb(filename, wb)


def set_cell(column, row, value, filename, sheetname):
    wb = load_wb(filename)
    ws = wb[sheetname]
    if column > 25:
        raise ValueError("Exceeds standard AZ column reference")
    column_letter = chr(ord('A') + column)
    ref = '{}{}'.format(column_letter, row)
    ws[ref] = value
    save_wb(filename, wb)


def new_wb():
    wb = Workbook()
    return wb


def filename_generate(dataframe):
    name = "".join(dataframe["Name"].split(" "))
    date = "-".join(dataframe["Checkout End"].split("/"))
    return "EquipmentChk_{}_{}.pdf".format(name, date)


def read_default_equipment(filename):
    # reads the Equipment tab to get default equipments for each lab week
    return None


def list_classes(sections_list):
    result = []
    for section in sections_list.keys():
        value = section.split("-")[0]
        if value not in result:
            result.append(value)
    return result


def wrap_cells(filename, sheetname):
    wb = load_wb(filename)
    ws = wb[sheetname]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = styles.Alignment(wrap_text=True)
    save_wb(filename, wb)
